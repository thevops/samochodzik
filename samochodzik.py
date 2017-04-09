# -*- coding: utf-8 -*-

# others
import sys
from datetime import datetime, timedelta, date
import calendar
import os
# QT
from PyQt4.QtCore import *
from PyQt4.QtGui import *
# PROJECT FILES
import models
import dialogs
# SQLAlchemy
from sqlalchemy import and_, func, desc
# save to Excel XLSX
from openpyxl import Workbook

# --------------------- GUI ---------------------------------------

class GUI(QTabWidget):
    def __init__(self, parent = None):
        super(GUI, self).__init__(parent)
        self.resize(1050, 600)
        #self.move(300, 250)
        #self.setFixedSize(750, 550)
        self.center()
        self.setWindowTitle("Samochodzik")
        self.setWindowIcon(QIcon("icons/main.png"))

        self.tab_podsumowanie = QWidget()
        self.tab_paliwo = QWidget()
        self.tab_naprawa = QWidget()
        self.tab_wymiana = QWidget()
        self.tab_przeglad = QWidget()
        self.tab_ubezpieczenie = QWidget()
        self.tab_raporty = QWidget()
        self.tab_samochody = QWidget()
        self.tab_oProgramie = QWidget()

        self.addTab(self.tab_podsumowanie, u"Podsumowanie")
        self.addTab(self.tab_paliwo,u"Paliwo")
        self.addTab(self.tab_naprawa,u"Naprawa")
        self.addTab(self.tab_wymiana,u"Wymiana o/f")
        self.addTab(self.tab_przeglad,u"Przegląd techniczny")
        self.addTab(self.tab_ubezpieczenie,u"Ubezpieczenie")
        self.addTab(self.tab_raporty,u"Raporty")
        self.addTab(self.tab_samochody,u"Pojazdy")
        self.addTab(self.tab_oProgramie,u"O programie")

        #self.tabBar().setTabTextColor(0, Qt.blue)
        self.tabBar().setStyleSheet('font-weight: bold;')
        self.tabBar().setTabIcon(0,QIcon("icons/podsumowanie.png"))
        self.tabBar().setTabIcon(1,QIcon("icons/fuel.png"))
        self.tabBar().setTabIcon(2,QIcon("icons/repair.png"))
        self.tabBar().setTabIcon(3,QIcon("icons/replace.png"))
        self.tabBar().setTabIcon(4,QIcon("icons/technical.png"))
        self.tabBar().setTabIcon(5,QIcon("icons/insurance.png"))
        self.tabBar().setTabIcon(6,QIcon("icons/raport.png"))
        self.tabBar().setTabIcon(7,QIcon("icons/vehicle.png"))
        self.tabBar().setTabIcon(8,QIcon("icons/about.png"))

        # DATABASE
        self.session = models.database_connect()

        # Przypomnienia
        self.ubezpieczenie_sie_zbliza()
        self.przeglad_sie_zbliza()
        self.wymiana_sie_zbliza()

        self.tab_podsumowanieUI()
        self.tab_paliwoUI()
        self.tab_naprawaUI()
        self.tab_wymianaUI()
        self.tab_przegladUI()
        self.tab_ubezpieczenieUI()
        self.tab_raportyUI()
        self.tab_samochodyUI()
        self.tab_oProgramieUI()

        self.currentChanged.connect(self.refresh_tabs)


    def center(self):
        qr = self.frameGeometry()
        cp = QDesktopWidget().availableGeometry().center()
        qr.moveCenter(cp)
        self.move(qr.topLeft())

    # -------------------------------------------------------     FUNCKJE POMOCNICZE
    def years_list(self):  # tworzy liste lat
        year = datetime.today().year
        years = QStringList()
        for i in range(5):
            years.append(QString(str(year-i)))
        return years

    def monthName2int(self, month):
        names = ["styczeń","luty","marzec","kwiecień","maj","czerwiec","lipiec","sierpień","wrzesień","październik","listopad","grudzień"]
        return names.index(month)+1

    def polish_months(self):
        return [u"styczeń",u"luty",u"marzec",u"kwiecień",u"maj",u"czerwiec",u"lipiec",u"sierpień",
                                                u"wrzesień",u"październik",u"listopad",u"grudzień"]

    def getPath2Save(self):
        dialog = QFileDialog()
        dialog.setFileMode(QFileDialog.Directory)
        dialog.setOption(QFileDialog.ShowDirsOnly)
        if dialog.exec_():
            sciezka_zapisu = dialog.selectedFiles()[0]
            return sciezka_zapisu


    def refresh_samochod(self):  # odswieza QComboBox'y po dodaniu nowego samochodu
        self.paliwo_auto.clear()
        self.naprawa_auto.clear()
        self.wymiana_auto.clear()
        self.przeglad_auto.clear()
        self.ubezpieczenie_auto.clear()
        for i in self.session.query(models.Car).all():
            self.paliwo_auto.addItem(i.name, i.id)
            self.naprawa_auto.addItem(i.name, i.id)
            self.wymiana_auto.addItem(i.name, i.id)
            self.przeglad_auto.addItem(i.name, i.id)
            self.ubezpieczenie_auto.addItem(i.name, i.id)

    def ubezpieczenie_sie_zbliza(self):
        przypomnienie = []
        for i in self.session.query(models.Car).all():
            # down - nie jestem piewien. Powinno zwrocic ostatnie ubezpieczenie samochodu
            last_ubezp = self.session.query(models.Insurance).filter_by(car=i).order_by(desc(models.Insurance.id)).first()
            if last_ubezp:
                # ubezpieczenie mija w ciagu 7 dni lub jest przeterminowane
                if last_ubezp.date_to - timedelta(days=7) <= date.today(): #and last_ubezp.date_to >= date.today():
                    przypomnienie.append(last_ubezp)
        if przypomnienie:
            do_ubezp = [i.car.name for i in przypomnienie]
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle(u"Ubezpieczenie")
            msg.setText(u"Ubezpiecz pojazdy:\n" + u'\n'.join(do_ubezp))
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

    def przeglad_sie_zbliza(self):
        przypomnienie = []
        for i in self.session.query(models.Car).all():
            # down - nie jestem piewien. Powinno zwrocic ostatnie przeglad samochodu
            last_techrev = self.session.query(models.TechnicalReview).filter_by(car=i).order_by(desc(models.TechnicalReview.id)).first()
            if last_techrev:
                # przeglad mija w ciagu 7 dni lub jest przeterminowany
                if last_techrev.techrev_next_date - timedelta(days=7) <= date.today(): #and last_techrev.techrev_next_date >= date.today():
                    przypomnienie.append(last_techrev)
        if przypomnienie:
            do_przegladu = [i.car.name for i in przypomnienie]
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle(u"Przegląd techniczny")
            msg.setText(u"Wykonaj przegląd techniczny pojazdów:\n" + u'\n'.join(do_przegladu))
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

    def wymiana_sie_zbliza(self):
        przypomnienie = []
        for i in self.session.query(models.Car).all():
            # down - nie jestem piewien. Powinno zwrocic ostatnie przeglad samochodu
            last_replecement = self.session.query(models.Replecement).filter_by(car=i).order_by(desc(models.Replecement.id)).first()
            if last_replecement:
                # przeglad mija w ciagu 7 dni lub jest przeterminowany
                if last_replecement.replace_date_next - timedelta(days=7) <= date.today():
                    przypomnienie.append(last_replecement)
        if przypomnienie:
            do_przegladu = [i.car.name for i in przypomnienie]
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle(u"Wymiana olejów/filtrów")
            msg.setText(u"Wykonaj wymianę olejów/filtrów w pojazdach:\n" + u'\n'.join(do_przegladu))
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

    def podsumowanie_ubezpieczenie_func(self):
        auta_ubezp = []
        for i in self.session.query(models.Car).all():
            tmp = self.session.query(models.Insurance).filter_by(car=i).order_by(desc(models.Insurance.date_to)).first()
            if tmp != None:
                auta_ubezp.append(tmp)

        if len(auta_ubezp) > 0:
            self.podsumowanie_ubezpieczenie.setColumnCount(5)
            self.podsumowanie_ubezpieczenie.setEditTriggers(QAbstractItemView.NoEditTriggers)
            self.podsumowanie_ubezpieczenie.setHorizontalHeaderLabels((u"Pojazd","Data od","Data do","Ubezpieczyciel", "Typ"))#,"Uwagi"))
            self.podsumowanie_ubezpieczenie.verticalHeader().setVisible(False)
            self.podsumowanie_ubezpieczenie.setRowCount(len(auta_ubezp))
            for i, col in enumerate(auta_ubezp):
                self.podsumowanie_ubezpieczenie.setItem(i, 0, QTableWidgetItem(u'{}'.format(col.car.name)))
                self.podsumowanie_ubezpieczenie.setItem(i, 1, QTableWidgetItem(u'{}'.format(col.date_from)))
                self.podsumowanie_ubezpieczenie.setItem(i, 2, QTableWidgetItem(u'{}'.format(col.date_to)))
                self.podsumowanie_ubezpieczenie.setItem(i, 3, QTableWidgetItem(u'{}'.format(col.firm)))
                #self.podsumowanie_ubezpieczenie.setItem(i, 4, QTableWidgetItem(u'{}'.format(col.price)))
                self.podsumowanie_ubezpieczenie.setItem(i, 4, QTableWidgetItem(u'{}'.format(col.type_of)))
                #self.podsumowanie_ubezpieczenie.setItem(i, 6, QTableWidgetItem(u'{}'.format(col.comments)))
                for p in range(5):
                    self.podsumowanie_ubezpieczenie.item(i,p).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            #self.podsumowanie_ubezpieczenie.resizeRowsToContents()
            self.podsumowanie_ubezpieczenie.setColumnWidth(0,120)
            self.podsumowanie_ubezpieczenie.setColumnWidth(1,90)
            self.podsumowanie_ubezpieczenie.setColumnWidth(2,90)
            self.podsumowanie_ubezpieczenie.setColumnWidth(3,150)
            self.podsumowanie_ubezpieczenie.horizontalHeader().setStretchLastSection(True)

    def podsumowanie_przeglad_func(self):
        auta_przegl = []
        for i in self.session.query(models.Car).all():
            tmp = self.session.query(models.TechnicalReview).filter_by(car=i).order_by(desc(models.TechnicalReview.techrev_next_date)).first()
            if tmp != None:
                auta_przegl.append(tmp)

        if len(auta_przegl) > 0:
            self.podsumowanie_przeglad.setColumnCount(4)
            self.podsumowanie_przeglad.setEditTriggers(QAbstractItemView.NoEditTriggers)
            self.podsumowanie_przeglad.setHorizontalHeaderLabels((u"Pojazd",u"Data przeglądu",u"Następny przegląd", u"Wartość [zł]"))#,"Uwagi"))
            self.podsumowanie_przeglad.verticalHeader().setVisible(False)
            self.podsumowanie_przeglad.setRowCount(len(auta_przegl))
            for i, col in enumerate(auta_przegl):
                self.podsumowanie_przeglad.setItem(i, 0, QTableWidgetItem(u'{}'.format(col.car.name)))
                self.podsumowanie_przeglad.setItem(i, 1, QTableWidgetItem(u'{}'.format(col.techrev_date)))
                self.podsumowanie_przeglad.setItem(i, 2, QTableWidgetItem(u'{}'.format(col.techrev_next_date)))
                self.podsumowanie_przeglad.setItem(i, 3, QTableWidgetItem(u'{}'.format(col.price)))
                #self.podsumowanie_przeglad.setItem(i, 4, QTableWidgetItem(u'{}'.format(col.comments)))
                for p in range(4):
                    self.podsumowanie_przeglad.item(i,p).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            #self.podsumowanie_przeglad.resizeRowsToContents()
            self.podsumowanie_przeglad.setColumnWidth(0,130)
            self.podsumowanie_przeglad.setColumnWidth(1,140)
            self.podsumowanie_przeglad.setColumnWidth(2,140)
            self.podsumowanie_przeglad.horizontalHeader().setStretchLastSection(True)

    def podsumowanie_wymiana_func(self):
        auta_wym = []
        for i in self.session.query(models.Car).all():
            tmp = self.session.query(models.Replecement).filter_by(car=i).order_by(desc(models.Replecement.replace_date_next)).first()
            if tmp != None:
                auta_wym.append(tmp)

        if len(auta_wym) > 0:
            self.podsumowanie_wymiana.setColumnCount(4)
            self.podsumowanie_wymiana.setEditTriggers(QAbstractItemView.NoEditTriggers)
            self.podsumowanie_wymiana.setHorizontalHeaderLabels((u"Pojazd","Data wymiany",u"Następna wymiana",u"Wartość [zł]"))#, "Opis"))
            self.podsumowanie_wymiana.verticalHeader().setVisible(False)
            self.podsumowanie_wymiana.setRowCount(len(auta_wym))
            for i, col in enumerate(auta_wym):
                self.podsumowanie_wymiana.setItem(i, 0, QTableWidgetItem(u'{}'.format(col.car.name)))
                self.podsumowanie_wymiana.setItem(i, 1, QTableWidgetItem(u'{}'.format(col.replace_date)))
                self.podsumowanie_wymiana.setItem(i, 2, QTableWidgetItem(u'{}'.format(col.replace_date_next)))
                self.podsumowanie_wymiana.setItem(i, 3, QTableWidgetItem(u'{}'.format(col.price)))
                #self.podsumowanie_wymiana.setItem(i, 4, QTableWidgetItem(u'{}'.format(col.description)))
                for p in range(4):
                    self.podsumowanie_wymiana.item(i,p).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            #self.podsumowanie_wymiana.resizeRowsToContents()
            self.podsumowanie_wymiana.setColumnWidth(0,130)
            self.podsumowanie_wymiana.setColumnWidth(1,140)
            self.podsumowanie_wymiana.setColumnWidth(2,140)
            self.podsumowanie_wymiana.horizontalHeader().setStretchLastSection(True)

    # --------------------------------------------------------------------

    def refresh_tabs(self):  #, i):
    	self.podsumowanie_ubezpieczenie_func()
    	self.podsumowanie_przeglad_func()
    	self.podsumowanie_wymiana_func()
    	#print i # printuje index tabsa
        
      
    def tab_podsumowanieUI(self):
        grid = QGridLayout()
        grid.setSpacing(5)
        grid.setContentsMargins(5,10,5,5)
        
        grid.addWidget(QLabel(u"Ostatnie ubezpieczenia"), 0,0)
        self.podsumowanie_ubezpieczenie = QTableWidget()
        self.podsumowanie_ubezpieczenie_func()
        grid.addWidget(self.podsumowanie_ubezpieczenie, 1,0,1,2)

        obrazek = QLabel()
        obrazek.setScaledContents(True)
        obrazek.setFixedSize(450,250)
        obrazek.setPixmap(QPixmap(os.getcwd() + "/icons/car.png"))
        grid.addWidget(obrazek,1,2,1,2)

        grid.addWidget(QLabel(u"Ostatnie przeglądy"), 2,0)
        self.podsumowanie_przeglad = QTableWidget()
        self.podsumowanie_przeglad_func()
        grid.addWidget(self.podsumowanie_przeglad, 3,0,1,2)

        grid.addWidget(QLabel(u"Ostatnie wymiany"), 2,2)
        self.podsumowanie_wymiana = QTableWidget()
        self.podsumowanie_wymiana_func()
        grid.addWidget(self.podsumowanie_wymiana, 3,2,1,2)

        self.tab_podsumowanie.setLayout(grid)


    def tab_paliwoUI(self):
        grid = QGridLayout()
        grid.setSpacing(5)
        grid.setContentsMargins(5,5,5,5)
        #grid.setColumnStretch(0,1)
        
        label1 = QLabel("Pojazd")
        label1.setFixedSize(40,25)
        label1.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        grid.addWidget(label1, 0, 0)

        self.paliwo_auto = QComboBox()
        for i in self.session.query(models.Car).all():
            self.paliwo_auto.addItem(i.name, i.id)
        self.paliwo_auto.setMinimumHeight(25)
        grid.addWidget(self.paliwo_auto,0,1, 1,2)

        label2 = QLabel("Od")
        label2.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        grid.addWidget(label2, 1, 0)

        self.paliwo_data_od_rok = QComboBox()
        self.paliwo_data_od_rok.addItems(self.years_list())
        self.paliwo_data_od_rok.setFixedSize(70,25)
        grid.addWidget(self.paliwo_data_od_rok,1,2)

        self.paliwo_data_od_miesiac = QComboBox()
        self.paliwo_data_od_miesiac.addItems(self.polish_months())
        self.paliwo_data_od_miesiac.setCurrentIndex(datetime.today().month-1)
        self.paliwo_data_od_miesiac.setFixedSize(90,25)
        grid.addWidget(self.paliwo_data_od_miesiac,1,1)

        label3 = QLabel("Do")
        label3.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        grid.addWidget(label3, 2, 0)

        self.paliwo_data_do_rok = QComboBox()
        self.paliwo_data_do_rok.addItems(self.years_list())
        self.paliwo_data_do_rok.setFixedSize(70,25)
        grid.addWidget(self.paliwo_data_do_rok,2,2)

        self.paliwo_data_do_miesiac = QComboBox()
        self.paliwo_data_do_miesiac.addItems(self.polish_months())
        self.paliwo_data_do_miesiac.setCurrentIndex(datetime.today().month-1)
        self.paliwo_data_do_miesiac.setFixedSize(90,25)
        grid.addWidget(self.paliwo_data_do_miesiac,2,1)
   
        przerwa = QLabel("")
        przerwa.setFixedSize(50,25)
        grid.addWidget(przerwa, 0, 3,3,1)

        self.paliwo_filtruj_btn = QPushButton(u"Odśwież")
        self.paliwo_filtruj_btn.setFixedSize(100,25)
        grid.addWidget(self.paliwo_filtruj_btn, 0, 4, 3, 1)
        self.paliwo_filtruj_btn.clicked.connect(self.click_paliwo_filtruj)

        self.paliwo_dodaj_btn = QPushButton("Dodaj")
        self.paliwo_dodaj_btn.setFixedSize(100,25)
        grid.addWidget(self.paliwo_dodaj_btn, 0, 5, 3, 1)
        self.paliwo_dodaj_btn.clicked.connect(self.click_paliwo_dodaj)

        self.paliwo_edytuj_btn = QPushButton("Edytuj")
        self.paliwo_edytuj_btn.setFixedSize(100,25)
        grid.addWidget(self.paliwo_edytuj_btn, 0, 6, 3, 1)
        self.paliwo_edytuj_btn.clicked.connect(self.click_paliwo_edytuj)

        self.paliwo_eksportuj_btn = QPushButton("Eksportuj")
        self.paliwo_eksportuj_btn.setFixedSize(100,25)
        grid.addWidget(self.paliwo_eksportuj_btn, 0, 7, 3, 1)
        self.paliwo_eksportuj_btn.clicked.connect(self.click_paliwo_eksportuj)
        self.sql_samochod_fuels = None

        self.paliwo_tablewidget = QTableWidget()
        grid.addWidget(self.paliwo_tablewidget, 3, 0, 1, 9)

        self.tab_paliwo.setLayout(grid)
      
    def tab_naprawaUI(self):
        grid = QGridLayout()
        grid.setSpacing(5)
        grid.setContentsMargins(5,5,5,5)
        
        label1 = QLabel("Pojazd")
        label1.setFixedSize(40,25)
        label1.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        grid.addWidget(label1, 0, 0)

        self.naprawa_auto = QComboBox()
        for i in self.session.query(models.Car).all():
            self.naprawa_auto.addItem(i.name, i.id)
        self.naprawa_auto.setMinimumHeight(25)
        grid.addWidget(self.naprawa_auto,0,1, 1,2)

        label2 = QLabel("Od")
        label2.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        grid.addWidget(label2, 1, 0)

        self.naprawa_data_od_rok = QComboBox()
        self.naprawa_data_od_rok.addItems(self.years_list())
        self.naprawa_data_od_rok.setFixedSize(70,25)
        grid.addWidget(self.naprawa_data_od_rok,1,2)

        self.naprawa_data_od_miesiac = QComboBox()
        self.naprawa_data_od_miesiac.addItems(self.polish_months())
        self.naprawa_data_od_miesiac.setCurrentIndex(datetime.today().month-1)
        self.naprawa_data_od_miesiac.setFixedSize(90,25)
        grid.addWidget(self.naprawa_data_od_miesiac,1,1)

        label3 = QLabel("Do")
        label3.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        grid.addWidget(label3, 2, 0)

        self.naprawa_data_do_rok = QComboBox()
        self.naprawa_data_do_rok.addItems(self.years_list())
        self.naprawa_data_do_rok.setFixedSize(70,25)
        grid.addWidget(self.naprawa_data_do_rok,2,2)

        self.naprawa_data_do_miesiac = QComboBox()
        self.naprawa_data_do_miesiac.addItems(self.polish_months())
        self.naprawa_data_do_miesiac.setCurrentIndex(datetime.today().month-1)
        self.naprawa_data_do_miesiac.setFixedSize(90,25)
        grid.addWidget(self.naprawa_data_do_miesiac,2,1)
   

        przerwa = QLabel("")
        przerwa.setFixedSize(50,25)
        grid.addWidget(przerwa, 0, 3,3,1)

        self.naprawa_filtruj_btn = QPushButton(u"Odśwież")
        self.naprawa_filtruj_btn.setFixedSize(100,25)
        grid.addWidget(self.naprawa_filtruj_btn, 0, 4, 3, 1)
        self.naprawa_filtruj_btn.clicked.connect(self.click_naprawa_filtruj)

        self.naprawa_dodaj_btn = QPushButton("Dodaj")
        self.naprawa_dodaj_btn.setFixedSize(100,25)
        grid.addWidget(self.naprawa_dodaj_btn, 0, 5, 3, 1)
        self.naprawa_dodaj_btn.clicked.connect(self.click_naprawa_dodaj)

        self.naprawa_edytuj_btn = QPushButton("Edytuj")
        self.naprawa_edytuj_btn.setFixedSize(100,25)
        grid.addWidget(self.naprawa_edytuj_btn, 0, 6, 3, 1)
        self.naprawa_edytuj_btn.clicked.connect(self.click_naprawa_edytuj)

        self.naprawa_eksportuj_btn = QPushButton("Eksportuj")
        self.naprawa_eksportuj_btn.setFixedSize(100,25)
        grid.addWidget(self.naprawa_eksportuj_btn, 0, 7, 3, 1)
        self.naprawa_eksportuj_btn.clicked.connect(self.click_naprawa_eksportuj)
        self.sql_samochod_reparations = None

        self.naprawa_tablewidget = QTableWidget()
        grid.addWidget(self.naprawa_tablewidget, 3, 0, 1, 9)

        self.tab_naprawa.setLayout(grid)

      
    def tab_wymianaUI(self):
        grid = QGridLayout()
        grid.setSpacing(5)
        grid.setContentsMargins(5,5,5,5)
        
        label1 = QLabel("Pojazd")
        label1.setFixedSize(40,25)
        label1.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        grid.addWidget(label1, 0, 0)

        self.wymiana_auto = QComboBox()
        for i in self.session.query(models.Car).all():
            self.wymiana_auto.addItem(i.name, i.id)
        self.wymiana_auto.setMinimumHeight(25)
        grid.addWidget(self.wymiana_auto,0,1, 1,2)

        label2 = QLabel("Od")
        label2.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        grid.addWidget(label2, 1, 0)

        self.wymiana_data_od_rok = QComboBox()
        self.wymiana_data_od_rok.addItems(self.years_list())
        self.wymiana_data_od_rok.setFixedSize(70,25)
        grid.addWidget(self.wymiana_data_od_rok,1,2)

        self.wymiana_data_od_miesiac = QComboBox()
        self.wymiana_data_od_miesiac.addItems(self.polish_months())
        self.wymiana_data_od_miesiac.setCurrentIndex(datetime.today().month-1)
        self.wymiana_data_od_miesiac.setFixedSize(90,25)
        grid.addWidget(self.wymiana_data_od_miesiac,1,1)

        label3 = QLabel("Do")
        label3.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        grid.addWidget(label3, 2, 0)

        self.wymiana_data_do_rok = QComboBox()
        self.wymiana_data_do_rok.addItems(self.years_list())
        self.wymiana_data_do_rok.setFixedSize(70,25)
        grid.addWidget(self.wymiana_data_do_rok,2,2)

        self.wymiana_data_do_miesiac = QComboBox()
        self.wymiana_data_do_miesiac.addItems(self.polish_months())
        self.wymiana_data_do_miesiac.setCurrentIndex(datetime.today().month-1)
        self.wymiana_data_do_miesiac.setFixedSize(90,25)
        grid.addWidget(self.wymiana_data_do_miesiac,2,1)
   
        przerwa = QLabel("")
        przerwa.setFixedSize(50,25)
        grid.addWidget(przerwa, 0, 3,3,1)

        self.wymiana_filtruj_btn = QPushButton(u"Odśwież")
        self.wymiana_filtruj_btn.setFixedSize(100,25)
        grid.addWidget(self.wymiana_filtruj_btn, 0, 4, 3, 1)
        self.wymiana_filtruj_btn.clicked.connect(self.click_wymiana_filtruj)

        self.wymiana_dodaj_btn = QPushButton("Dodaj")
        self.wymiana_dodaj_btn.setFixedSize(100,25)
        grid.addWidget(self.wymiana_dodaj_btn, 0, 5, 3, 1)
        self.wymiana_dodaj_btn.clicked.connect(self.click_wymiana_dodaj)

        self.wymiana_edytuj_btn = QPushButton("Edytuj")
        self.wymiana_edytuj_btn.setFixedSize(100,25)
        grid.addWidget(self.wymiana_edytuj_btn, 0, 6, 3, 1)
        self.wymiana_edytuj_btn.clicked.connect(self.click_wymiana_edytuj)

        self.wymiana_eksportuj_btn = QPushButton("Eksportuj")
        self.wymiana_eksportuj_btn.setFixedSize(100,25)
        grid.addWidget(self.wymiana_eksportuj_btn, 0, 7, 3, 1)
        self.wymiana_eksportuj_btn.clicked.connect(self.click_wymiana_eksportuj)
        self.sql_samochod_replecements = None

        self.wymiana_tablewidget = QTableWidget()
        grid.addWidget(self.wymiana_tablewidget, 3, 0, 1, 9)

        self.tab_wymiana.setLayout(grid)

    def tab_przegladUI(self):
        grid = QGridLayout()
        grid.setSpacing(10)
        grid.setContentsMargins(5,15,5,5)
        
        label1 = QLabel("Pojazd")
        label1.setFixedSize(40,25)
        label1.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        grid.addWidget(label1, 0, 0)

        self.przeglad_auto = QComboBox()
        for i in self.session.query(models.Car).all():
            self.przeglad_auto.addItem(i.name, i.id)
        self.przeglad_auto.setFixedSize(200,25)
        grid.addWidget(self.przeglad_auto,0,1)
   
        self.przeglad_filtruj_btn = QPushButton(u"Odśwież")
        self.przeglad_filtruj_btn.setFixedSize(100,25)
        grid.addWidget(self.przeglad_filtruj_btn, 0, 2)
        self.przeglad_filtruj_btn.clicked.connect(self.click_przeglad_filtruj)

        self.przeglad_dodaj_btn = QPushButton("Dodaj")
        self.przeglad_dodaj_btn.setFixedSize(100,25)
        grid.addWidget(self.przeglad_dodaj_btn, 0, 3)
        self.przeglad_dodaj_btn.clicked.connect(self.click_przeglad_dodaj)

        self.przeglad_edytuj_btn = QPushButton("Edytuj")
        self.przeglad_edytuj_btn.setFixedSize(100,25)
        grid.addWidget(self.przeglad_edytuj_btn, 0, 4)
        self.przeglad_edytuj_btn.clicked.connect(self.click_przeglad_edytuj)

        self.przeglad_eksportuj_btn = QPushButton("Eksportuj")
        self.przeglad_eksportuj_btn.setFixedSize(100,25)
        grid.addWidget(self.przeglad_eksportuj_btn, 0, 5)
        self.przeglad_eksportuj_btn.clicked.connect(self.click_przeglad_eksportuj)
        self.sql_samochod_techreviews = None

        self.przeglad_tablewidget = QTableWidget()
        grid.addWidget(self.przeglad_tablewidget, 1, 0, 1, 7)

        self.tab_przeglad.setLayout(grid)

    def tab_ubezpieczenieUI(self):
        grid = QGridLayout()
        grid.setSpacing(10)
        grid.setContentsMargins(5,15,5,5)
        
        label1 = QLabel("Pojazd")
        label1.setFixedSize(40,25)
        label1.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        grid.addWidget(label1, 0, 0)

        self.ubezpieczenie_auto = QComboBox()
        for i in self.session.query(models.Car).all():
            self.ubezpieczenie_auto.addItem(i.name, i.id)
        self.ubezpieczenie_auto.setFixedSize(200,25)
        grid.addWidget(self.ubezpieczenie_auto,0,1)
   
        self.ubezpieczenie_filtruj_btn = QPushButton(u"Odśwież")
        self.ubezpieczenie_filtruj_btn.setFixedSize(100,25)
        grid.addWidget(self.ubezpieczenie_filtruj_btn, 0, 2)
        self.ubezpieczenie_filtruj_btn.clicked.connect(self.click_ubezpieczenie_filtruj)

        self.ubezpieczenie_dodaj_btn = QPushButton("Dodaj")
        self.ubezpieczenie_dodaj_btn.setFixedSize(100,25)
        grid.addWidget(self.ubezpieczenie_dodaj_btn, 0, 3)
        self.ubezpieczenie_dodaj_btn.clicked.connect(self.click_ubezpieczenie_dodaj)

        self.ubezpieczenie_edytuj_btn = QPushButton("Edytuj")
        self.ubezpieczenie_edytuj_btn.setFixedSize(100,25)
        grid.addWidget(self.ubezpieczenie_edytuj_btn, 0, 4)
        self.ubezpieczenie_edytuj_btn.clicked.connect(self.click_ubezpieczenie_edytuj)

        self.ubezpieczenie_eksportuj_btn = QPushButton("Eksportuj")
        self.ubezpieczenie_eksportuj_btn.setFixedSize(100,25)
        grid.addWidget(self.ubezpieczenie_eksportuj_btn, 0, 5)
        self.ubezpieczenie_eksportuj_btn.clicked.connect(self.click_ubezpieczenie_eksportuj)
        self.sql_samochod_insurance = None

        self.ubezpieczenie_tablewidget = QTableWidget()
        grid.addWidget(self.ubezpieczenie_tablewidget, 1, 0, 1, 7)

        self.tab_ubezpieczenie.setLayout(grid)

    def tab_raportyUI(self):
        grid = QGridLayout()
        grid.setSpacing(5)
        grid.setContentsMargins(5,5,5,5)

        l0 = QLabel("Data")
        l0.setFixedSize(200,25)
        l0.setAlignment(Qt.AlignCenter)
        grid.addWidget(l0, 0, 1,1,2)

        l1 = QLabel("Od")
        l1.setFixedSize(30,25)
        l1.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        grid.addWidget(l1, 1, 0)

        self.raporty_data_od_miesiac = QComboBox()
        self.raporty_data_od_miesiac.addItems(self.polish_months())
        self.raporty_data_od_miesiac.setCurrentIndex(datetime.today().month-1)
        self.raporty_data_od_miesiac.setFixedSize(100,25)
        grid.addWidget(self.raporty_data_od_miesiac,1,1)

        self.raporty_data_od_rok = QComboBox()
        self.raporty_data_od_rok.addItems(self.years_list())
        self.raporty_data_od_rok.setFixedSize(70,25)
        grid.addWidget(self.raporty_data_od_rok,1,2)      

        l2 = QLabel("Do")
        l2.setFixedSize(30,25)
        l2.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        grid.addWidget(l2, 2, 0)

        self.raporty_data_do_miesiac = QComboBox()
        self.raporty_data_do_miesiac.addItems(self.polish_months())
        self.raporty_data_do_miesiac.setCurrentIndex(datetime.today().month-1)
        self.raporty_data_do_miesiac.setFixedSize(100,25)
        grid.addWidget(self.raporty_data_do_miesiac,2,1)

        self.raporty_data_do_rok = QComboBox()
        self.raporty_data_do_rok.addItems(self.years_list())
        self.raporty_data_do_rok.setFixedSize(70,25)
        grid.addWidget(self.raporty_data_do_rok,2,2)

        l3 = QLabel("Kategoria")
        l3.setFixedSize(200,25)
        l3.setAlignment(Qt.AlignCenter)
        grid.addWidget(l3, 0, 3,1,2)

        self.raporty_kategoria = QComboBox()
        self.raporty_kategoria.addItem("Paliwo",0)
        self.raporty_kategoria.addItem("Naprawa",1)
        self.raporty_kategoria.addItem("Wymiana o/f",2)
        self.raporty_kategoria.addItem(u"Przegląd techniczny",3)
        self.raporty_kategoria.addItem("Ubezpieczenie",4)
        self.raporty_kategoria.setMinimumHeight(25)
        grid.addWidget(self.raporty_kategoria,1,3,1,2)

        self.raporty_generuj_btn = QPushButton(u"Odśwież")
        self.raporty_generuj_btn.setFixedSize(100,25)
        grid.addWidget(self.raporty_generuj_btn, 2, 3)
        self.raporty_generuj_btn.clicked.connect(self.click_raporty_generuj)

        self.raporty_eksportuj_btn = QPushButton(u"Eksportuj")
        self.raporty_eksportuj_btn.setFixedSize(100,25)
        grid.addWidget(self.raporty_eksportuj_btn, 2, 4)
        self.raporty_eksportuj_btn.clicked.connect(self.click_raporty_eksportuj)
        
        tmp = QLabel("") #  sztuczna przerwa
        tmp.setMinimumWidth(300)
        grid.addWidget(tmp,0,5)

        grid.addWidget(QLabel(u'Suma "Wartości"'),2,6)
        #suma
        self.suma = QLineEdit()
        self.suma.setReadOnly(True)
        self.suma.setMaximumWidth(100)
        grid.addWidget(self.suma,2,7)

        self.raporty_tablewidget = QTableWidget()
        grid.addWidget(self.raporty_tablewidget, 3,0,1,8)
        #grid.setColumnStretch(4,4)
        #grid.setRowStretch(4,4)

        


        self.tab_raporty.setLayout(grid)

    def tab_samochodyUI(self):
        grid = QGridLayout()
        grid.setSpacing(5)
        grid.setContentsMargins(10,20,10,20)
        
   
        self.samochody_odswiez_btn = QPushButton(u"Odśwież")
        self.samochody_odswiez_btn.setFixedSize(100,25)
        grid.addWidget(self.samochody_odswiez_btn, 0, 0)
        self.samochody_odswiez_btn.clicked.connect(self.click_samochody_odswiez)

        self.samochody_dodaj_btn = QPushButton("Dodaj")
        self.samochody_dodaj_btn.setFixedSize(100,25)
        grid.addWidget(self.samochody_dodaj_btn, 0, 1)
        self.samochody_dodaj_btn.clicked.connect(self.click_samochody_dodaj)

        self.samochody_edytuj_btn = QPushButton("Edytuj")
        self.samochody_edytuj_btn.setFixedSize(100,25)
        grid.addWidget(self.samochody_edytuj_btn, 0, 2)
        self.samochody_edytuj_btn.clicked.connect(self.click_samochody_edytuj)

        self.sql_samochody = None

        self.samochody_tablewidget = QTableWidget()
        grid.addWidget(self.samochody_tablewidget, 1, 0, 1, 4)

        self.tab_samochody.setLayout(grid)

    def tab_oProgramieUI(self):
        grid = QGridLayout()
        tekst = QTextEdit()
        tekst.setMaximumHeight(400)
        tekst.setReadOnly(True)
        tekst.textCursor().insertHtml(u"""
                                        <font size="6">
                                        <table>
                                        <tr>
                                         <td>Autor:</td><td>Krzysztof Łuczak</td>
                                        </tr>
                                        <tr>
                                         <td>E-mail:</td><td>krzysztof.luczak99@gmail.com</td>
                                        </tr>
                                        <tr>
                                         <td>Github:</td><td>https://github.com/vizarch</td>
                                        </tr>
                                        </table>
                                        <br /><br /><br /><br /><br /><br /><br />
                                        <center>
                                        Copyright © 2017, Krzysztof Łuczak
                                        </center></font>
                                        """)
        grid.addWidget(tekst,0,0)
        self.tab_oProgramie.setLayout(grid)


    # ----------------------   EVENTS   -----------------------------
    # ----------------------   PALIWO   -----------------------------

    def click_paliwo_filtruj(self):
        id_samochod = self.paliwo_auto.itemData(self.paliwo_auto.currentIndex()).toPyObject()
        sql_samochod = self.session.query(models.Car).get(id_samochod)

        data_od = [int(self.paliwo_data_od_rok.currentText()), self.monthName2int(self.paliwo_data_od_miesiac.currentText().toUtf8())]
        data_do = [int(self.paliwo_data_do_rok.currentText()), self.monthName2int(self.paliwo_data_do_miesiac.currentText().toUtf8())]
        start_date = date(data_od[0], data_od[1], 1)
        num_days = calendar.monthrange(data_do[0], data_do[1])[1]
        end_date = date(data_do[0], data_do[1], num_days)

        self.sql_samochod_fuels = self.session.query(models.Fuel).filter_by(car=sql_samochod).filter(
                                            and_(models.Fuel.refuel_date>=start_date, models.Fuel.refuel_date<=end_date)).order_by(
                                            models.Fuel.refuel_date).all()

        self.paliwo_tablewidget.setColumnCount(7)
        self.paliwo_tablewidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        #self.paliwo_tablewidget.setColumnWidth(0,50)
        self.paliwo_tablewidget.horizontalHeader().setStretchLastSection(True)
        self.paliwo_tablewidget.setHorizontalHeaderLabels(("ID", u"Pojazd","Data",u"Ilość",u"Cena [zł/litr]",u"Wartość [zł]","Przebieg"))
        self.paliwo_tablewidget.verticalHeader().setVisible(False)
        self.paliwo_tablewidget.setRowCount(len(self.sql_samochod_fuels))
        for i, col in enumerate(self.sql_samochod_fuels):
            self.paliwo_tablewidget.setItem(i, 0, QTableWidgetItem(u'{}'.format(col.id)))
            self.paliwo_tablewidget.setItem(i, 1, QTableWidgetItem(u'{}'.format(col.car.name)))
            self.paliwo_tablewidget.setItem(i, 2, QTableWidgetItem(u'{}'.format(col.refuel_date)))
            self.paliwo_tablewidget.setItem(i, 3, QTableWidgetItem(u'{}'.format(col.quantity)))
            self.paliwo_tablewidget.setItem(i, 4, QTableWidgetItem(u'{}'.format(col.price)))
            self.paliwo_tablewidget.setItem(i, 5, QTableWidgetItem(u'{}'.format(col.value)))
            self.paliwo_tablewidget.setItem(i, 6, QTableWidgetItem(u'{}'.format(col.mileage)))
            for p in range(7):
                self.paliwo_tablewidget.item(i,p).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
        self.paliwo_tablewidget.resizeRowsToContents()
        #self.paliwo_tablewidget.resizeColumnsToContents()


    def click_paliwo_dodaj(self):
        self.dialog_paliwo_dodaj = dialogs.DG_paliwo_dodaj(self.session, self)
        if self.dialog_paliwo_dodaj.exec_():
            id_samochod = self.dialog_paliwo_dodaj.auto.itemData(self.dialog_paliwo_dodaj.auto.currentIndex()).toPyObject()
            sql_samochod = self.session.query(models.Car).get(id_samochod)
            data = self.dialog_paliwo_dodaj.data.selectedDate().toPyDate()
            ilosc = self.dialog_paliwo_dodaj.ilosc.value()
            cena = self.dialog_paliwo_dodaj.cena.value()
            #wartosc = self.dialog_paliwo_dodaj.wartosc.value()
            wartosc = ilosc*cena
            przebieg = self.dialog_paliwo_dodaj.przebieg.value()
            nowe_tankowanie = models.Fuel(auto_id=sql_samochod.id, refuel_date=data, quantity=ilosc, price=cena, mileage=przebieg, value=wartosc)
            self.session.add(nowe_tankowanie)
            self.session.commit()

    def click_paliwo_edytuj(self):
        if not self.sql_samochod_fuels:
            return
        self.dialog_paliwo_edytuj = dialogs.DG_paliwo_edytuj(self.session, self)
        if self.dialog_paliwo_edytuj.exec_():
            id_tankowania = int(self.dialog_paliwo_edytuj.id.currentText())
            data = self.dialog_paliwo_edytuj.data.selectedDate().toPyDate()
            ilosc = self.dialog_paliwo_edytuj.ilosc.value()
            cena = self.dialog_paliwo_edytuj.cena.value()
            #wartosc = self.dialog_paliwo_edytuj.wartosc.value()
            wartosc = ilosc*cena
            przebieg = self.dialog_paliwo_edytuj.przebieg.value()
            self.session.query(models.Fuel).filter_by(id=id_tankowania).update({"refuel_date": data, "quantity": ilosc,
                                                                                                "price":cena,"mileage":przebieg,
                                                                                                "value":wartosc})
            self.session.commit()

    def click_paliwo_eksportuj(self):
        path = self.getPath2Save()
        if self.sql_samochod_fuels and path:
            excel = Workbook(write_only=True)
            strona = excel.create_sheet()
            strona.append(["Data",u"Ilość [litry]", u"Cena [zł/litr]",u"Wartość [zł]", "Przebieg [km]"])
            for row in self.sql_samochod_fuels:
                strona.append([str(row.refuel_date),float(row.quantity),float(row.price), float(row.value),float(row.mileage)])
            nazwa_pliku = "paliwo_"+self.sql_samochod_fuels[0].car.name+".xlsx"
            sciezka_zapisu = os.path.join(str(path),nazwa_pliku)
            excel.save(sciezka_zapisu)

            # komunikat
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle(u"Operacja udana")
            msg.setText(u'Dane zostały zapisane w pliku: '+ nazwa_pliku)
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
        else:
            # komunikat
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle(u"Błąd")
            msg.setText(u'Najpierw wczytaj dane poprzez przycisk "Odśwież"')
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()


    # --------------------   NAPRAWA   -------------------------
    def click_naprawa_filtruj(self):
        id_samochod = self.naprawa_auto.itemData(self.naprawa_auto.currentIndex()).toPyObject()
        sql_samochod = self.session.query(models.Car).get(id_samochod)

        data_od = [int(self.naprawa_data_od_rok.currentText()), self.monthName2int(self.naprawa_data_od_miesiac.currentText().toUtf8())]
        data_do = [int(self.naprawa_data_do_rok.currentText()), self.monthName2int(self.naprawa_data_do_miesiac.currentText().toUtf8())]
        start_date = date(data_od[0], data_od[1], 1)
        num_days = calendar.monthrange(data_do[0], data_do[1])[1]
        end_date = date(data_do[0], data_do[1], num_days)

        self.sql_samochod_reparations = self.session.query(models.Reparation).filter_by(car=sql_samochod).filter(
                                            and_(models.Reparation.repair_date>=start_date, models.Reparation.repair_date<=end_date)).order_by(
                                            models.Reparation.repair_date).all()

        self.naprawa_tablewidget.setColumnCount(5)
        #self.naprawa_tablewidget.setColumnWidth(0,50)
        self.naprawa_tablewidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.naprawa_tablewidget.horizontalHeader().setStretchLastSection(True)
        self.naprawa_tablewidget.setHorizontalHeaderLabels(("ID", u"Pojazd","Data",u"Wartość [zł]", "Opis"))
        self.naprawa_tablewidget.verticalHeader().setVisible(False)
        self.naprawa_tablewidget.setRowCount(len(self.sql_samochod_reparations))
        for i, col in enumerate(self.sql_samochod_reparations):
            self.naprawa_tablewidget.setItem(i, 0, QTableWidgetItem(u'{}'.format(col.id)))
            self.naprawa_tablewidget.setItem(i, 1, QTableWidgetItem(u'{}'.format(col.car.name)))
            self.naprawa_tablewidget.setItem(i, 2, QTableWidgetItem(u'{}'.format(col.repair_date)))
            self.naprawa_tablewidget.setItem(i, 3, QTableWidgetItem(u'{}'.format(col.price)))
            self.naprawa_tablewidget.setItem(i, 4, QTableWidgetItem(u'{}'.format(col.description)))
            for p in range(5):
                self.naprawa_tablewidget.item(i,p).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
        #self.naprawa_tablewidget.resizeColumnsToContents()
        self.naprawa_tablewidget.resizeRowsToContents()


    def click_naprawa_dodaj(self):
        self.dialog_naprawa_dodaj = dialogs.DG_naprawa_dodaj(self.session, self)
        if self.dialog_naprawa_dodaj.exec_():
            id_samochod = self.dialog_naprawa_dodaj.auto.itemData(self.dialog_naprawa_dodaj.auto.currentIndex()).toPyObject()
            sql_samochod = self.session.query(models.Car).get(id_samochod)
            data = self.dialog_naprawa_dodaj.data.selectedDate().toPyDate()
            cena = self.dialog_naprawa_dodaj.cena.value()
            opis = u'{}'.format(self.dialog_naprawa_dodaj.opis.toPlainText())  # unicode
            nowa_naprawa = models.Reparation(auto_id=sql_samochod.id, description=opis, repair_date=data, price=cena)
            self.session.add(nowa_naprawa)
            self.session.commit()

    def click_naprawa_edytuj(self):
        if not self.sql_samochod_reparations:
            return
        self.dialog_naprawa_edytuj = dialogs.DG_naprawa_edytuj(self.session, self)
        if self.dialog_naprawa_edytuj.exec_():
            id_naprawy = int(self.dialog_naprawa_edytuj.id.currentText())
            data = self.dialog_naprawa_edytuj.data.selectedDate().toPyDate()
            cena = self.dialog_naprawa_edytuj.cena.value()
            opis = u'{}'.format(self.dialog_naprawa_edytuj.opis.toPlainText())  # unicode
            self.session.query(models.Reparation).filter_by(id=id_naprawy).update({"repair_date": data,
                                                                                                "price":cena,"description":opis})
            self.session.commit()

    def click_naprawa_eksportuj(self):
        path = self.getPath2Save()
        if self.sql_samochod_reparations and path:
            excel = Workbook(write_only=True)
            strona = excel.create_sheet()
            strona.append(["Data",u"Wartość [zł]", "Opis"])
            for row in self.sql_samochod_reparations:
                strona.append([str(row.repair_date),float(row.price), u'{}'.format(row.description)])
            nazwa_pliku = "naprawa_"+self.sql_samochod_reparations[0].car.name+".xlsx"
            sciezka_zapisu = os.path.join(str(path),nazwa_pliku)
            excel.save(sciezka_zapisu)

            # komunikat
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle(u"Operacja udana")
            msg.setText(u'Dane zostały zapisane w pliku: '+ nazwa_pliku)
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
        else:
            # komunikat
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle(u"Błąd")
            msg.setText(u'Najpierw wczytaj dane poprzez przycisk "Odśwież"')
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

    # ------------------------------  WYMIANA  --------------------------

    def click_wymiana_filtruj(self):
        id_samochod = self.wymiana_auto.itemData(self.wymiana_auto.currentIndex()).toPyObject()
        sql_samochod = self.session.query(models.Car).get(id_samochod)

        data_od = [int(self.wymiana_data_od_rok.currentText()), self.monthName2int(self.wymiana_data_od_miesiac.currentText().toUtf8())]
        data_do = [int(self.wymiana_data_do_rok.currentText()), self.monthName2int(self.wymiana_data_do_miesiac.currentText().toUtf8())]
        start_date = date(data_od[0], data_od[1], 1)
        num_days = calendar.monthrange(data_do[0], data_do[1])[1]
        end_date = date(data_do[0], data_do[1], num_days)

        self.sql_samochod_replecements = self.session.query(models.Replecement).filter_by(car=sql_samochod).filter(
                                            and_(models.Replecement.replace_date>=start_date, models.Replecement.replace_date<=end_date)).order_by(
                                            models.Replecement.replace_date).all()
        

        self.wymiana_tablewidget.setColumnCount(7)
        self.wymiana_tablewidget.setColumnWidth(3,130)
        self.wymiana_tablewidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.wymiana_tablewidget.horizontalHeader().setStretchLastSection(True)
        self.wymiana_tablewidget.setHorizontalHeaderLabels(("ID", u"Pojazd","Data wymiany",u"Następna wymiana",u"Wartość [zł]", "Przebieg", "Opis"))
        self.wymiana_tablewidget.verticalHeader().setVisible(False)
        self.wymiana_tablewidget.setRowCount(len(self.sql_samochod_replecements))
        for i, col in enumerate(self.sql_samochod_replecements):
            self.wymiana_tablewidget.setItem(i, 0, QTableWidgetItem(u'{}'.format(col.id)))
            self.wymiana_tablewidget.setItem(i, 1, QTableWidgetItem(u'{}'.format(col.car.name)))
            self.wymiana_tablewidget.setItem(i, 2, QTableWidgetItem(u'{}'.format(col.replace_date)))
            self.wymiana_tablewidget.setItem(i, 3, QTableWidgetItem(u'{}'.format(col.replace_date_next)))
            self.wymiana_tablewidget.setItem(i, 4, QTableWidgetItem(u'{}'.format(col.price)))
            self.wymiana_tablewidget.setItem(i, 5, QTableWidgetItem(u'{}'.format(col.mileage)))
            self.wymiana_tablewidget.setItem(i, 6, QTableWidgetItem(u'{}'.format(col.description)))
            for p in range(7):
                self.wymiana_tablewidget.item(i,p).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
        #self.wymiana_tablewidget.resizeColumnsToContents()
        self.wymiana_tablewidget.resizeRowsToContents()



    def click_wymiana_dodaj(self):
        self.dialog_wymiana_dodaj = dialogs.DG_wymiana_dodaj(self.session, self)
        if self.dialog_wymiana_dodaj.exec_():
            id_samochod = self.dialog_wymiana_dodaj.auto.itemData(self.dialog_wymiana_dodaj.auto.currentIndex()).toPyObject()
            sql_samochod = self.session.query(models.Car).get(id_samochod)
            data = self.dialog_wymiana_dodaj.data.selectedDate().toPyDate()
            data_next = self.dialog_wymiana_dodaj.data_next.selectedDate().toPyDate()
            cena = self.dialog_wymiana_dodaj.cena.value()
            przebieg = self.dialog_wymiana_dodaj.przebieg.value()
            opis = u'{}'.format(self.dialog_wymiana_dodaj.opis.toPlainText())  # unicode
            nowa_wymiana = models.Replecement(auto_id=sql_samochod.id, description=opis, replace_date=data, price=cena, mileage=przebieg, replace_date_next=data_next)
            self.session.add(nowa_wymiana)
            self.session.commit()

    def click_wymiana_edytuj(self):
        if not self.sql_samochod_replecements:
            return
        self.dialog_wymiana_edytuj = dialogs.DG_wymiana_edytuj(self.session, self)
        if self.dialog_wymiana_edytuj.exec_():
            id_wymiana = int(self.dialog_wymiana_edytuj.id.currentText())
            data = self.dialog_wymiana_edytuj.data.selectedDate().toPyDate()
            data_next = self.dialog_wymiana_edytuj.data_next.selectedDate().toPyDate()
            cena = self.dialog_wymiana_edytuj.cena.value()
            przebieg = self.dialog_wymiana_edytuj.przebieg.value()
            opis = u'{}'.format(self.dialog_wymiana_edytuj.opis.toPlainText())  # unicode
            self.session.query(models.Replecement).filter_by(id=id_wymiana).update({"replace_date": data,
                                                                                                "price":cena,"description":opis,
                                                                                                "mileage": przebieg,"replace_date_next":data_next})
            self.session.commit()

    def click_wymiana_eksportuj(self):
        path = self.getPath2Save()
        if self.sql_samochod_replecements and path:
            excel = Workbook(write_only=True)
            strona = excel.create_sheet()
            strona.append(["Data wymiany",u"Następna wymiana",u"Wartość [zł]", "Opis", "Przebieg"])
            for row in self.sql_samochod_replecements:
                strona.append([str(row.replace_date),str(row.replace_date_next),float(row.price), u'{}'.format(row.description), float(row.mileage)])
            nazwa_pliku = "wymiana_"+self.sql_samochod_replecements[0].car.name+".xlsx"
            sciezka_zapisu = os.path.join(str(path),nazwa_pliku)
            excel.save(sciezka_zapisu)

            # komunikat
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle(u"Operacja udana")
            msg.setText(u'Dane zostały zapisane w pliku: '+ nazwa_pliku)
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
        else:
            # komunikat
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle(u"Błąd")
            msg.setText(u'Najpierw wczytaj dane poprzez przycisk "Odśwież"')
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

    # -----------------    PRZEGLAD   ------------------------------
    def click_przeglad_filtruj(self):
        id_samochod = self.przeglad_auto.itemData(self.przeglad_auto.currentIndex()).toPyObject()
        sql_samochod = self.session.query(models.Car).get(id_samochod)


        self.sql_samochod_techreviews = self.session.query(models.TechnicalReview).filter_by(car=sql_samochod).order_by(
                                                models.TechnicalReview.techrev_date).all()
        
        self.przeglad_tablewidget.setColumnCount(6)
        self.przeglad_tablewidget.setColumnWidth(3,150)
        self.przeglad_tablewidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.przeglad_tablewidget.horizontalHeader().setStretchLastSection(True)
        self.przeglad_tablewidget.setHorizontalHeaderLabels(("ID", u"Pojazd", u"Data przeglądu", u"Następny przegląd", u"Wartość [zł]", u"Uwagi"))
        self.przeglad_tablewidget.verticalHeader().setVisible(False)
        self.przeglad_tablewidget.setRowCount(len(self.sql_samochod_techreviews))
        for i, col in enumerate(self.sql_samochod_techreviews):
            self.przeglad_tablewidget.setItem(i, 0, QTableWidgetItem(u'{}'.format(col.id)))
            self.przeglad_tablewidget.setItem(i, 1, QTableWidgetItem(u'{}'.format(col.car.name)))
            self.przeglad_tablewidget.setItem(i, 2, QTableWidgetItem(u'{}'.format(col.techrev_date)))
            self.przeglad_tablewidget.setItem(i, 3, QTableWidgetItem(u'{}'.format(col.techrev_next_date)))
            self.przeglad_tablewidget.setItem(i, 4, QTableWidgetItem(u'{}'.format(col.price)))
            self.przeglad_tablewidget.setItem(i, 5, QTableWidgetItem(u'{}'.format(col.comments)))
            for p in range(6):
                self.przeglad_tablewidget.item(i,p).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
        #self.przeglad_tablewidget.resizeColumnsToContents()
        self.przeglad_tablewidget.resizeRowsToContents()


    def click_przeglad_dodaj(self):
        self.dialog_przeglad_dodaj = dialogs.DG_przeglad_dodaj(self.session, self)
        if self.dialog_przeglad_dodaj.exec_():
            id_samochod = self.dialog_przeglad_dodaj.auto.itemData(self.dialog_przeglad_dodaj.auto.currentIndex()).toPyObject()
            sql_samochod = self.session.query(models.Car).get(id_samochod)
            data = self.dialog_przeglad_dodaj.data.selectedDate().toPyDate()
            data_next = self.dialog_przeglad_dodaj.data_next.selectedDate().toPyDate()
            cena = self.dialog_przeglad_dodaj.cena.value()
            uwagi = u'{}'.format(self.dialog_przeglad_dodaj.uwagi.toPlainText())
            nowa_przeglad = models.TechnicalReview(auto_id=sql_samochod.id, techrev_date=data, techrev_next_date=data_next, price=cena, comments=uwagi)
            self.session.add(nowa_przeglad)
            self.session.commit()

    def click_przeglad_edytuj(self):
        if not self.sql_samochod_techreviews:
            return
        self.dialog_przeglad_edytuj = dialogs.DG_przeglad_edytuj(self.session, self)
        if self.dialog_przeglad_edytuj.exec_():
            id_przeglad = int(self.dialog_przeglad_edytuj.id.currentText())
            data = self.dialog_przeglad_edytuj.data.selectedDate().toPyDate()
            data_next = self.dialog_przeglad_edytuj.data_next.selectedDate().toPyDate()
            cena = self.dialog_przeglad_edytuj.cena.value()
            uwagi = u'{}'.format(self.dialog_przeglad_edytuj.uwagi.toPlainText())
            self.session.query(models.TechnicalReview).filter_by(id=id_przeglad).update({"techrev_date": data,
                                                                                                "price":cena, 
                                                                                                "techrev_next_date": data_next, "comments":uwagi})
            self.session.commit()

    def click_przeglad_eksportuj(self):
        path = self.getPath2Save()
        if self.sql_samochod_techreviews and path:
            excel = Workbook(write_only=True)
            strona = excel.create_sheet()
            strona.append([u"Data przeglądu",u"Następny przegląd", u"Wartość [zł]",u"Uwagi"])
            for row in self.sql_samochod_techreviews:
                strona.append([str(row.techrev_date),str(row.techrev_next_date), float(row.price),str(row.comments)])
            nazwa_pliku = "przeglad_"+self.sql_samochod_techreviews[0].car.name+".xlsx"
            sciezka_zapisu = os.path.join(str(path),nazwa_pliku)
            excel.save(sciezka_zapisu)

            # komunikat
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle(u"Operacja udana")
            msg.setText(u'Dane zostały zapisane w pliku: '+ nazwa_pliku)
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
        else:
            # komunikat
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle(u"Błąd")
            msg.setText(u'Najpierw wczytaj dane poprzez przycisk "Odśwież"')
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

    # -----------------    UBEZPIECZENIE   ------------------------------
    def click_ubezpieczenie_filtruj(self):
        id_samochod = self.ubezpieczenie_auto.itemData(self.ubezpieczenie_auto.currentIndex()).toPyObject()
        sql_samochod = self.session.query(models.Car).get(id_samochod)


        self.sql_samochod_insurance = self.session.query(models.Insurance).filter_by(car=sql_samochod).order_by(
                                            models.Insurance.date_from).all()
        

        self.ubezpieczenie_tablewidget.setColumnCount(8)
        #self.ubezpieczenie_tablewidget.setColumnWidth(0,50)
        self.ubezpieczenie_tablewidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.ubezpieczenie_tablewidget.horizontalHeader().setStretchLastSection(True)
        self.ubezpieczenie_tablewidget.setHorizontalHeaderLabels(("ID", u"Pojazd","Data od","Data do", u"Wartość", "Ubezpieczyciel", "Typ","Uwagi"))
        self.ubezpieczenie_tablewidget.verticalHeader().setVisible(False)
        self.ubezpieczenie_tablewidget.setRowCount(len(self.sql_samochod_insurance))
        for i, col in enumerate(self.sql_samochod_insurance):
            self.ubezpieczenie_tablewidget.setItem(i, 0, QTableWidgetItem(u'{}'.format(col.id)))
            self.ubezpieczenie_tablewidget.setItem(i, 1, QTableWidgetItem(u'{}'.format(col.car.name)))
            self.ubezpieczenie_tablewidget.setItem(i, 2, QTableWidgetItem(u'{}'.format(col.date_from)))
            self.ubezpieczenie_tablewidget.setItem(i, 3, QTableWidgetItem(u'{}'.format(col.date_to)))
            self.ubezpieczenie_tablewidget.setItem(i, 4, QTableWidgetItem(u'{}'.format(col.price)))
            self.ubezpieczenie_tablewidget.setItem(i, 5, QTableWidgetItem(u'{}'.format(col.firm)))
            self.ubezpieczenie_tablewidget.setItem(i, 6, QTableWidgetItem(u'{}'.format(col.type_of)))
            self.ubezpieczenie_tablewidget.setItem(i, 7, QTableWidgetItem(u'{}'.format(col.comments)))
            for p in range(8):
                self.ubezpieczenie_tablewidget.item(i,p).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
        #self.ubezpieczenie_tablewidget.resizeColumnsToContents()
        self.ubezpieczenie_tablewidget.resizeRowsToContents()


    def click_ubezpieczenie_dodaj(self):
        self.dialog_ubezpieczenie_dodaj = dialogs.DG_ubezpieczenie_dodaj(self.session, self)
        if self.dialog_ubezpieczenie_dodaj.exec_():
            id_samochod = self.dialog_ubezpieczenie_dodaj.auto.itemData(self.dialog_ubezpieczenie_dodaj.auto.currentIndex()).toPyObject()
            sql_samochod = self.session.query(models.Car).get(id_samochod)
            data = self.dialog_ubezpieczenie_dodaj.data.selectedDate().toPyDate()
            data_next = self.dialog_ubezpieczenie_dodaj.data_next.selectedDate().toPyDate()
            cena = self.dialog_ubezpieczenie_dodaj.cena.value()
            firma = u'{}'.format(self.dialog_ubezpieczenie_dodaj.firma.toPlainText())
            typ = str(self.dialog_ubezpieczenie_dodaj.typ.currentText())
            uwagi = u'{}'.format(self.dialog_ubezpieczenie_dodaj.uwagi.toPlainText())
            nowa_ubezpieczenie = models.Insurance(auto_id=sql_samochod.id, date_from=data, date_to=data_next, price=cena, firm=firma,
                                                    type_of=typ, comments=uwagi)
            self.session.add(nowa_ubezpieczenie)
            self.session.commit()

    def click_ubezpieczenie_edytuj(self):
        if not self.sql_samochod_insurance:
            return
        self.dialog_ubezpieczenie_edytuj = dialogs.DG_ubezpieczenie_edytuj(self.session, self)
        if self.dialog_ubezpieczenie_edytuj.exec_():
            id_ubezpieczenie = int(self.dialog_ubezpieczenie_edytuj.id.currentText())
            data = self.dialog_ubezpieczenie_edytuj.data.selectedDate().toPyDate()
            data_next = self.dialog_ubezpieczenie_edytuj.data_next.selectedDate().toPyDate()
            cena = self.dialog_ubezpieczenie_edytuj.cena.value()
            firma = u'{}'.format(self.dialog_ubezpieczenie_edytuj.firma.toPlainText())
            typ = str(self.dialog_ubezpieczenie_edytuj.typ.currentText())
            uwagi = u'{}'.format(self.dialog_ubezpieczenie_dodaj.uwagi.toPlainText())
            self.session.query(models.Insurance).filter_by(id=id_ubezpieczenie).update({"date_from": data,
                                                                                                "price":cena, 
                                                                                                "date_to": data_next, "firm": firma,
                                                                                                "type_of":typ,"comments":uwagi})
            self.session.commit()

    def click_ubezpieczenie_eksportuj(self):
        path = self.getPath2Save()
        if self.sql_samochod_insurance and path:
            excel = Workbook(write_only=True)
            strona = excel.create_sheet()
            strona.append(["Data od ","Data do", u"Wartość [zł]", "Ubezpieczyciel", "Typ","Uwagi"])
            for row in self.sql_samochod_insurance:
                strona.append([str(row.date_from),str(row.date_to), float(row.price), str(row.firm), str(row.type_of),str(row.comments)])
            nazwa_pliku = "ubezpieczenie_"+self.sql_samochod_insurance[0].car.name+".xlsx"
            sciezka_zapisu = os.path.join(str(path),nazwa_pliku)
            excel.save(sciezka_zapisu)

            # komunikat
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Information)
            msg.setWindowTitle(u"Operacja udana")
            msg.setText(u'Dane zostały zapisane w pliku: '+ nazwa_pliku)
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()
        else:
            # komunikat
            msg = QMessageBox(self)
            msg.setIcon(QMessageBox.Critical)
            msg.setWindowTitle(u"Błąd")
            msg.setText(u'Najpierw wczytaj dane poprzez przycisk "Odśwież"')
            msg.setStandardButtons(QMessageBox.Ok)
            msg.exec_()

    # ---------------------   RAPORTY    ------------------------
    def click_raporty_eksportuj(self):
        path = self.getPath2Save()
        if path:
            kategoria = self.raporty_kategoria.itemData(self.raporty_kategoria.currentIndex()).toPyObject() # int - index
            data_od = [int(self.raporty_data_od_rok.currentText()), self.monthName2int(self.raporty_data_od_miesiac.currentText().toUtf8())]
            data_do = [int(self.raporty_data_do_rok.currentText()), self.monthName2int(self.raporty_data_do_miesiac.currentText().toUtf8())]
            start_date = date(data_od[0], data_od[1], 1)
            num_days = calendar.monthrange(data_do[0], data_do[1])[1]
            end_date = date(data_do[0], data_do[1], num_days)

            if kategoria == 0:
                res = self.session.query(models.Fuel).order_by(models.Fuel.refuel_date).filter(
                                            and_(models.Fuel.refuel_date>=start_date, models.Fuel.refuel_date<=end_date)).all()
                excel = Workbook(write_only=True)
                strona = excel.create_sheet()
                strona.append([u"Pojazd ","Data",u"Ilość [litr]", u"Cena [zł]", u"Wartość","Przebieg"])
                for row in res:
                    strona.append([str(row.car.name), str(row.refuel_date), float(row.quantity), float(row.price),float(row.value) ,float(row.mileage)])
                nazwa_pliku = "raport-Paliwo.xlsx"
                sciezka_zapisu = os.path.join(str(path),nazwa_pliku)
                excel.save(sciezka_zapisu)

                msg = QMessageBox(self)
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle(u"Operacja udana")
                msg.setText(u"Raport zapisany w pliku: "+nazwa_pliku)
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()

            if kategoria == 1:
                res = self.session.query(models.Reparation).order_by(models.Reparation.repair_date).filter(
                                            and_(models.Reparation.repair_date>=start_date, models.Reparation.repair_date<=end_date)).all()
                excel = Workbook(write_only=True)
                strona = excel.create_sheet()
                strona.append([u"Pojazd ","Data",u"Wartość [zł]", "Opis"])
                for row in res:
                    strona.append([str(row.car.name), str(row.repair_date), float(row.price), u'{}'.format(row.description)])
                nazwa_pliku = "raport-Naprawa.xlsx"
                sciezka_zapisu = os.path.join(str(path),nazwa_pliku)
                excel.save(sciezka_zapisu)

                msg = QMessageBox(self)
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle(u"Operacja udana")
                msg.setText(u"Raport zapisany w pliku: "+nazwa_pliku)
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()

            if kategoria == 2:
                res = self.session.query(models.Replecement).order_by(models.Replecement.replace_date).filter(
                                            and_(models.Replecement.replace_date>=start_date, models.Replecement.replace_date<=end_date)).all()
                excel = Workbook(write_only=True)
                strona = excel.create_sheet()
                strona.append([u"Pojazd ","Data wymiany",u"Następna wymiana",u"Wartość [zł]", "Opis"])
                for row in res:
                    strona.append([str(row.car.name), str(row.replace_date),str(row.replace_date_next), float(row.price), u'{}'.format(row.description)])
                nazwa_pliku = "raport-Wymiana.xlsx"
                sciezka_zapisu = os.path.join(str(path),nazwa_pliku)
                excel.save(sciezka_zapisu)

                msg = QMessageBox(self)
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle(u"Operacja udana")
                msg.setText(u"Raport zapisany w pliku: "+nazwa_pliku)
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()

            if kategoria == 3:
                res = self.session.query(models.TechnicalReview).order_by(models.TechnicalReview.techrev_date).filter(
                                            and_(models.TechnicalReview.techrev_date>=start_date, models.TechnicalReview.techrev_date<=end_date)).all()
                excel = Workbook(write_only=True)
                strona = excel.create_sheet()
                strona.append([u"Pojazd ",u"Data przeglądu",u"Następny przegląd", u"Wartość [zł]","Uwagi"])
                for row in res:
                    strona.append([str(row.car.name), str(row.techrev_date), str(row.techrev_next_date), float(row.price),str(row.comments)])
                nazwa_pliku = "raport-PrzegladTechniczny.xlsx"
                sciezka_zapisu = os.path.join(str(path),nazwa_pliku)
                excel.save(sciezka_zapisu)

                msg = QMessageBox(self)
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle(u"Operacja udana")
                msg.setText(u"Raport zapisany w pliku: "+nazwa_pliku)
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()

            if kategoria == 4:
                res = self.session.query(models.Insurance).order_by(models.Insurance.date_from).filter(
                                            and_(models.Insurance.date_from>=start_date, models.Insurance.date_from<=end_date)).all()
                excel = Workbook(write_only=True)
                strona = excel.create_sheet()
                strona.append([u"Pojazd ","Data od","Data do", u"Wartość [zł]", "Ubezpieczyciel", "Typ","Uwagi"])
                for row in res:
                    strona.append([str(row.car.name), str(row.date_from), str(row.date_to), float(row.price), str(row.firm), str(row.type_of),str(row.comments)])
                nazwa_pliku = "raport-Ubezpieczenie.xlsx"
                sciezka_zapisu = os.path.join(str(path),nazwa_pliku)
                excel.save(sciezka_zapisu)

                msg = QMessageBox(self)
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle(u"Operacja udana")
                msg.setText(u"Raport zapisany w pliku: "+nazwa_pliku)
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()

    def click_raporty_generuj(self):
        kategoria = self.raporty_kategoria.itemData(self.raporty_kategoria.currentIndex()).toPyObject() # int - index
        data_od = [int(self.raporty_data_od_rok.currentText()), self.monthName2int(self.raporty_data_od_miesiac.currentText().toUtf8())]
        data_do = [int(self.raporty_data_do_rok.currentText()), self.monthName2int(self.raporty_data_do_miesiac.currentText().toUtf8())]
        start_date = date(data_od[0], data_od[1], 1)
        num_days = calendar.monthrange(data_do[0], data_do[1])[1]
        end_date = date(data_do[0], data_do[1], num_days)

        suma = 0

        if kategoria == 0:
            res = self.session.query(models.Fuel).order_by(models.Fuel.refuel_date).filter(
                                        and_(models.Fuel.refuel_date>=start_date, models.Fuel.refuel_date<=end_date)).all()
            
            self.raporty_tablewidget.setColumnCount(6)
            self.raporty_tablewidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
            self.raporty_tablewidget.setHorizontalHeaderLabels((u"Pojazd","Data",u"Ilość",u"Cena [zł]",u"Wartość [zł]", "Przebieg"))
            self.raporty_tablewidget.verticalHeader().setVisible(False)
            self.raporty_tablewidget.setRowCount(len(res))
            for i, col in enumerate(res):
                self.raporty_tablewidget.setItem(i, 0, QTableWidgetItem(u'{}'.format(col.car.name)))
                self.raporty_tablewidget.setItem(i, 1, QTableWidgetItem(u'{}'.format(col.refuel_date)))
                self.raporty_tablewidget.setItem(i, 2, QTableWidgetItem(u'{}'.format(col.quantity)))
                self.raporty_tablewidget.setItem(i, 3, QTableWidgetItem(u'{}'.format(col.price)))
                suma+=col.value
                self.raporty_tablewidget.setItem(i, 4, QTableWidgetItem(u'{}'.format(col.value)))
                self.raporty_tablewidget.setItem(i, 5, QTableWidgetItem(u'{}'.format(col.mileage)))
                for p in range(6):
                    self.raporty_tablewidget.item(i,p).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.raporty_tablewidget.setColumnWidth(0,150)
            self.raporty_tablewidget.setColumnWidth(1,150)
            self.raporty_tablewidget.setColumnWidth(2,150)
            self.raporty_tablewidget.setColumnWidth(3,150)
            self.raporty_tablewidget.setColumnWidth(4,150)
            self.raporty_tablewidget.horizontalHeader().setStretchLastSection(True)

        if kategoria == 1:
            res = self.session.query(models.Reparation).order_by(models.Reparation.repair_date).filter(
                                        and_(models.Reparation.repair_date>=start_date, models.Reparation.repair_date<=end_date)).all()
            
            self.raporty_tablewidget.setColumnCount(4)
            self.raporty_tablewidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
            self.raporty_tablewidget.setHorizontalHeaderLabels((u"Pojazd","Data",u"Wartość [zł]", "Opis"))
            self.raporty_tablewidget.verticalHeader().setVisible(False)
            self.raporty_tablewidget.setRowCount(len(res))
            for i, col in enumerate(res):
                self.raporty_tablewidget.setItem(i, 0, QTableWidgetItem(u'{}'.format(col.car.name)))
                self.raporty_tablewidget.setItem(i, 1, QTableWidgetItem(u'{}'.format(col.repair_date)))
                self.raporty_tablewidget.setItem(i, 2, QTableWidgetItem(u'{}'.format(col.price)))
                suma+=col.price
                self.raporty_tablewidget.setItem(i, 3, QTableWidgetItem(u'{}'.format(col.description)))
                for p in range(4):
                    self.raporty_tablewidget.item(i,p).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.raporty_tablewidget.setColumnWidth(0,150)
            self.raporty_tablewidget.setColumnWidth(1,150)
            self.raporty_tablewidget.setColumnWidth(2,150)
            self.raporty_tablewidget.horizontalHeader().setStretchLastSection(True)

        if kategoria == 2:
            res = self.session.query(models.Replecement).order_by(models.Replecement.replace_date).filter(
                                        and_(models.Replecement.replace_date>=start_date, models.Replecement.replace_date<=end_date)).all()
            
            self.raporty_tablewidget.setColumnCount(5)
            self.raporty_tablewidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
            self.raporty_tablewidget.setHorizontalHeaderLabels((u"Pojazd","Data wymiany",u"Następna wymiana",u"Wartość [zł]", "Opis"))
            self.raporty_tablewidget.verticalHeader().setVisible(False)
            self.raporty_tablewidget.setRowCount(len(res))
            for i, col in enumerate(res):
                self.raporty_tablewidget.setItem(i, 0, QTableWidgetItem(u'{}'.format(col.car.name)))
                self.raporty_tablewidget.setItem(i, 1, QTableWidgetItem(u'{}'.format(col.replace_date)))
                self.raporty_tablewidget.setItem(i, 2, QTableWidgetItem(u'{}'.format(col.replace_date_next)))
                self.raporty_tablewidget.setItem(i, 3, QTableWidgetItem(u'{}'.format(col.price)))
                suma+=col.price
                self.raporty_tablewidget.setItem(i, 4, QTableWidgetItem(u'{}'.format(col.description)))
                for p in range(5):
                    self.raporty_tablewidget.item(i,p).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.raporty_tablewidget.setColumnWidth(0,150)
            self.raporty_tablewidget.setColumnWidth(1,150)
            self.raporty_tablewidget.setColumnWidth(2,150)
            self.raporty_tablewidget.setColumnWidth(3,150)
            self.raporty_tablewidget.horizontalHeader().setStretchLastSection(True)   

        if kategoria == 3:
            res = self.session.query(models.TechnicalReview).order_by(models.TechnicalReview.techrev_date).filter(
                                        and_(models.TechnicalReview.techrev_date>=start_date, models.TechnicalReview.techrev_date<=end_date)).all()
            
            self.raporty_tablewidget.setColumnCount(5)
            self.raporty_tablewidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
            self.raporty_tablewidget.setHorizontalHeaderLabels((u"Pojazd",u"Data przeglądu",u"Następny przegląd", u"Wartość [zł]","Uwagi"))
            self.raporty_tablewidget.verticalHeader().setVisible(False)
            self.raporty_tablewidget.setRowCount(len(res))
            self.raporty_tablewidget.setColumnWidth(2,150)
            for i, col in enumerate(res):
                self.raporty_tablewidget.setItem(i, 0, QTableWidgetItem(u'{}'.format(col.car.name)))
                self.raporty_tablewidget.setItem(i, 1, QTableWidgetItem(u'{}'.format(col.techrev_date)))
                self.raporty_tablewidget.setItem(i, 2, QTableWidgetItem(u'{}'.format(col.techrev_next_date)))
                self.raporty_tablewidget.setItem(i, 3, QTableWidgetItem(u'{}'.format(col.price)))
                suma+=col.price
                self.raporty_tablewidget.setItem(i, 4, QTableWidgetItem(u'{}'.format(col.comments)))
                for p in range(5):
                    self.raporty_tablewidget.item(i,p).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.raporty_tablewidget.setColumnWidth(0,150)
            self.raporty_tablewidget.setColumnWidth(1,150)
            self.raporty_tablewidget.setColumnWidth(2,150)
            self.raporty_tablewidget.setColumnWidth(3,100)
            self.raporty_tablewidget.horizontalHeader().setStretchLastSection(True)

        if kategoria == 4:
            res = self.session.query(models.Insurance).order_by(models.Insurance.date_from).filter(
                                        and_(models.Insurance.date_from>=start_date, models.Insurance.date_from<=end_date)).all()
                
            self.raporty_tablewidget.setColumnCount(6)
            self.raporty_tablewidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
            self.raporty_tablewidget.setHorizontalHeaderLabels((u"Pojazd","Data od","Data do", u"Wartość [zł]", "Typ","Uwagi"))
            self.raporty_tablewidget.verticalHeader().setVisible(False)
            self.raporty_tablewidget.setRowCount(len(res))
            for i, col in enumerate(res):
                self.raporty_tablewidget.setItem(i, 0, QTableWidgetItem(u'{}'.format(col.car.name)))
                self.raporty_tablewidget.setItem(i, 1, QTableWidgetItem(u'{}'.format(col.date_from)))
                self.raporty_tablewidget.setItem(i, 2, QTableWidgetItem(u'{}'.format(col.date_to)))
                self.raporty_tablewidget.setItem(i, 3, QTableWidgetItem(u'{}'.format(col.price)))
                suma+=col.price
                self.raporty_tablewidget.setItem(i, 4, QTableWidgetItem(u'{}'.format(col.type_of)))
                self.raporty_tablewidget.setItem(i, 5, QTableWidgetItem(u'{}'.format(col.comments)))
                for p in range(6):
                    self.raporty_tablewidget.item(i,p).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.raporty_tablewidget.setColumnWidth(0,130)
            self.raporty_tablewidget.setColumnWidth(1,130)
            self.raporty_tablewidget.setColumnWidth(2,130)
            self.raporty_tablewidget.setColumnWidth(3,100)
            self.raporty_tablewidget.setColumnWidth(4,70)
            self.raporty_tablewidget.horizontalHeader().setStretchLastSection(True)

        self.suma.setText(str(suma))
        self.raporty_tablewidget.resizeRowsToContents()


    # --------------------------------------   SAMOCHODY    -------------------------
    def click_samochody_odswiez(self):
        self.sql_samochody = self.session.query(models.Car).all()
        
        self.samochody_tablewidget.setColumnCount(5)
        self.samochody_tablewidget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        #self.samochody_tablewidget.horizontalHeader().setStretchLastSection(True)
        self.samochody_tablewidget.setHorizontalHeaderLabels(("ID", u"Nazwa","Numer rejestracji","VIN", "Opis"))
        self.samochody_tablewidget.verticalHeader().setVisible(False)
        self.samochody_tablewidget.setRowCount(len(self.sql_samochody))
        for i, col in enumerate(self.sql_samochody):
            self.samochody_tablewidget.setItem(i, 0, QTableWidgetItem(u'{}'.format(col.id)))
            self.samochody_tablewidget.setItem(i, 1, QTableWidgetItem(u'{}'.format(col.name)))
            self.samochody_tablewidget.setItem(i, 2, QTableWidgetItem(u'{}'.format(col.registration_num)))
            self.samochody_tablewidget.setItem(i, 3, QTableWidgetItem(u'{}'.format(col.vin)))
            self.samochody_tablewidget.setItem(i, 4, QTableWidgetItem(u'{}'.format(col.full_name)))
            for p in range(5):
                    self.samochody_tablewidget.item(i,p).setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
        #self.samochody_tablewidget.resizeColumnsToContents()
        self.samochody_tablewidget.horizontalHeader().setStretchLastSection(True)

    def click_samochody_dodaj(self):
        samochody_dialog = dialogs.DG_samochody_dodaj(self.session, self)
        if samochody_dialog.exec_():
            auto_nazwa = u'{}'.format(samochody_dialog.auto_nazwa.text())
            auto_pelnanazwa = u'{}'.format(samochody_dialog.auto_pelnanazwa.text())
            auto_rejestracja = u'{}'.format(samochody_dialog.auto_rejestracja.text())
            auto_vin = u'{}'.format(samochody_dialog.auto_vin.text())
            self.session.add(models.Car(auto_nazwa, auto_pelnanazwa, auto_rejestracja, auto_vin))
            self.session.commit()
        self.refresh_samochod()
        

    def click_samochody_edytuj(self):
        samochody_dialog = dialogs.DG_samochody_edytuj(self.session, self)
        if samochody_dialog.exec_():
            id_auta = int(samochody_dialog.id.currentText())
            auto_nazwa = u'{}'.format(samochody_dialog.auto_nazwa.text())
            auto_pelnanazwa = u'{}'.format(samochody_dialog.auto_pelnanazwa.text())
            auto_rejestracja = u'{}'.format(samochody_dialog.auto_rejestracja.text())
            auto_vin = u'{}'.format(samochody_dialog.auto_vin.text())

            self.session.query(models.Car).filter_by(id=id_auta).update({"name": auto_nazwa,"full_name":auto_pelnanazwa,
                                                                            "registration_num": auto_rejestracja, "vin":auto_vin})
            self.session.commit()
        self.refresh_samochod()


def main():
    app = QApplication(sys.argv)
    ex = GUI()
    ex.show()
    sys.exit(app.exec_())
   
if __name__ == '__main__':
    main()